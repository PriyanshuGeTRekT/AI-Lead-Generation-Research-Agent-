"""Build the Delhi-NCR IT/BPO/consulting call sheet (formatted, sales fill-in columns)."""
import json
import sqlite3

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DB = "data/warehouse.db"
OUT = "data/Delhi_NCR_IT_BPO_Consulting_CallSheet.xlsx"
I = ("IT services", "BPO", "consulting")

# palette (warm paper + charcoal + gold, matching the app)
INK = "2B2622"
GOLD = "C9A227"
PAPER = "FAF6EF"
BAND = "F0E9DC"
HDR = "3A332C"
GREEN = "E4EFE0"

con = sqlite3.connect(DB)
con.row_factory = sqlite3.Row
rows = [dict(r) for r in con.execute(
    f"""SELECT * FROM leads WHERE state='Delhi NCR' AND industry IN ({','.join('?'*3)})
        AND phone!='' AND status NOT IN ('excluded','disqualified','raw')""", I).fetchall()]
con.close()


def pl(r):
    try:
        return json.loads(r["payload"]) if r["payload"] else {}
    except Exception:
        return {}


def why(p):
    rs = p.get("signal_reasons") or []
    if isinstance(rs, str):
        try:
            rs = json.loads(rs)
        except Exception:
            rs = [rs]
    return "; ".join(rs[:3]) if rs else (", ".join(p.get("pain_points", [])[:2]) if p.get("pain_points") else "")


def fmt_phone(ph):
    d = "".join(ch for ch in str(ph) if ch.isdigit())
    if len(d) == 10:
        return f"+91 {d[:5]} {d[5:]}"
    return str(ph)


# rank: named first, then by signal score desc
for r in rows:
    r["_p"] = pl(r)
    r["_named"] = 1 if (r.get("dm_name") or "").strip() else 0
    r["_sig"] = r.get("signal_score") or 0
rows.sort(key=lambda r: (-r["_named"], -r["_sig"], r["company_name"].lower()))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Call Sheet"

HEAD = ["#", "Company", "Industry", "Decision-Maker", "Role", "Phone", "Phone source",
        "Email", "Website", "Location", "Signal", "Why now (HRMS angle)",
        "Call date", "Reached? (Y/N)", "Spoke to", "Has HRMS? (Y/N)",
        "Interest (Hot/Warm/Cold)", "Next step", "Notes"]
widths = [4, 34, 12, 20, 18, 17, 11, 26, 30, 16, 7, 40, 11, 13, 18, 15, 18, 24, 30]

thin = Side(style="thin", color="D8CEBC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor=HDR)
fill_band = PatternFill("solid", fgColor=BAND)
fill_paper = PatternFill("solid", fgColor=PAPER)
fill_entry = PatternFill("solid", fgColor=GREEN)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# title row
ws.merge_cells("A1:S1")
t = ws["A1"]
t.value = f"RazorInfotech — HRMS Leads · Delhi NCR × IT / BPO / Consulting · {len(rows)} verified-phone leads"
t.font = Font(name="Calibri", bold=True, size=14, color=INK)
t.fill = PatternFill("solid", fgColor=GOLD)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

# header
for ci, h in enumerate(HEAD, 1):
    cell = ws.cell(row=2, column=ci, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = border
ws.row_dimensions[2].height = 34

entry_cols = set(range(13, 20))  # sales fill-in columns
for i, r in enumerate(rows, 1):
    p = r["_p"]
    rr = i + 2
    vals = [
        i, r["company_name"], r["industry"], r.get("dm_name") or "",
        (p.get("dm_role") or "") if r.get("dm_name") else "",
        fmt_phone(r["phone"]), p.get("phone_source") or "",
        r.get("dm_email") or (p.get("contact_emails") or [""])[0] if isinstance(p.get("contact_emails"), list) else (r.get("dm_email") or ""),
        r.get("website") or "", r.get("location") or "",
        int(r["_sig"]), why(p),
        "", "", "", "", "", "", "",
    ]
    band = fill_band if i % 2 == 0 else fill_paper
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=rr, column=ci, value=v)
        cell.border = border
        cell.alignment = center if ci in (1, 3, 7, 11) else left
        cell.fill = fill_entry if ci in entry_cols else band
        if ci == 4 and v:
            cell.font = Font(bold=True, color=INK)
    ws.row_dimensions[rr].height = 30

for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:S{len(rows)+2}"

# ── Summary sheet ──
s = wb.create_sheet("Summary")
named = sum(r["_named"] for r in rows)
hot = sum(1 for r in rows if r["_sig"] >= 60)
by_ind = {}
for r in rows:
    by_ind[r["industry"]] = by_ind.get(r["industry"], 0) + 1
s["A1"] = "Campaign Summary"
s["A1"].font = Font(bold=True, size=14, color=INK)
summ = [
    ("Total verified-phone leads", len(rows)),
    ("With named decision-maker", f"{named}  ({round(named*100/len(rows))}%)"),
    ("Hot (signal score ≥ 60)", hot),
    ("Segment", "Delhi NCR × IT / BPO / Consulting"),
    ("Phone", "Verified business line (Google Maps / OSM)"),
    ("Names from", "Company's own website, read by LLM (free)"),
    ("", ""),
    ("HOW TO USE", ""),
    ("1.", "Work top-down — named + high-signal leads are first."),
    ("2.", "For unnamed leads, ask: 'May I speak with the owner / whoever handles HR & payroll?'"),
    ("3.", "Fill the green columns on every call — that data tells us what's converting."),
    ("4.", "Mark 'Has HRMS? = Y' to drop them; that's our disqualifier."),
]
for i, (k, v) in enumerate(summ, 3):
    s.cell(row=i, column=1, value=k).font = Font(bold=k in ("HOW TO USE",) or k.endswith("."), color=INK)
    s.cell(row=i, column=2, value=v)
s.column_dimensions["A"].width = 30
s.column_dimensions["B"].width = 52

bi = wb.create_sheet("By Industry")
bi["A1"] = "Leads by Industry"
bi["A1"].font = Font(bold=True, size=13, color=INK)
bi.cell(row=2, column=1, value="Industry").font = Font(bold=True)
bi.cell(row=2, column=2, value="Leads").font = Font(bold=True)
for i, (k, v) in enumerate(sorted(by_ind.items(), key=lambda x: -x[1]), 3):
    bi.cell(row=i, column=1, value=k)
    bi.cell(row=i, column=2, value=v)
bi.column_dimensions["A"].width = 20
bi.column_dimensions["B"].width = 10

wb.save(OUT)
print(f"Saved {OUT}")
print(f"  {len(rows)} leads | {named} named ({round(named*100/len(rows))}%) | {hot} hot (sig>=60)")
print(f"  by industry: {by_ind}")
