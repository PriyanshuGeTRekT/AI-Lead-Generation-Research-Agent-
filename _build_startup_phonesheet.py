"""Excel call sheet of harvested Startup-India leads (Delhi-NCR x DPIIT x Scaling)
that currently have a phone number. Enriched fields + sales fill-in columns."""
import sqlite3
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DB = "data/warehouse.db"
OUT = "data/StartupIndia_DelhiNCR_DPIIT_Scaling_PHONES.xlsx"
NCR = "(state='Delhi' OR lower(city) IN ('gurugram','gurgaon','noida','greater noida','gautam buddha nagar','ghaziabad','faridabad'))"
INK, GOLD, HDR, PAPER, BAND, GREEN = "2B2622", "C9A227", "3A332C", "FAF6EF", "F0E9DC", "E4EFE0"

con = sqlite3.connect(DB); con.row_factory = sqlite3.Row
rows = [dict(r) for r in con.execute(
    f"""SELECT company_name AS name, city, industry, stage, phone, dm_name, dm_role,
        linkedin, email, reg_email, website, cin, contact_status FROM (
          SELECT name AS company_name, city, industry, stage, phone, dm_name, dm_role,
                 linkedin, email, reg_email, website, cin, contact_status
          FROM startups WHERE {NCR} AND dpiit_certified=1 AND stage='Scaling'
            AND phone!='' AND phone IS NOT NULL)
        ORDER BY (CASE WHEN dm_name!='' AND dm_name IS NOT NULL THEN 0 ELSE 1 END),
                 (CASE WHEN linkedin!='' AND linkedin IS NOT NULL THEN 0 ELSE 1 END), company_name""").fetchall()]
con.close()

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Call Sheet"
HEAD = ["#", "Company", "City", "Industry", "Phone", "Decision-maker", "Role",
        "LinkedIn", "Email", "Website", "CIN",
        "Call date", "Reached? (Y/N)", "Spoke to", "Has HRMS? (Y/N)",
        "Interest (Hot/Warm/Cold)", "Next step", "Notes"]
widths = [4, 32, 14, 16, 17, 20, 20, 30, 26, 28, 23, 11, 13, 16, 15, 18, 22, 28]
thin = Side(style="thin", color="D8CEBC"); border = Border(thin, thin, thin, thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
entry_cols = set(range(12, 19))

ws.merge_cells("A1:R1"); t = ws["A1"]
t.value = f"RazorInfotech — Startup India · Delhi-NCR × DPIIT × Scaling · {len(rows)} leads WITH PHONE"
t.font = Font(bold=True, size=14, color=INK); t.fill = PatternFill("solid", fgColor=GOLD)
t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 26

for ci, h in enumerate(HEAD, 1):
    cell = ws.cell(row=2, column=ci, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11); cell.fill = PatternFill("solid", fgColor=HDR)
    cell.alignment = center; cell.border = border
ws.row_dimensions[2].height = 34

def em(r):
    return r.get("email") or r.get("reg_email") or ""

for i, r in enumerate(rows, 1):
    rr = i + 2
    vals = [i, r["name"], r.get("city") or "", r.get("industry") or "", r["phone"],
            r.get("dm_name") or "", r.get("dm_role") or "", r.get("linkedin") or "",
            em(r), r.get("website") or "", r.get("cin") or "", "", "", "", "", "", "", ""]
    band = PatternFill("solid", fgColor=BAND if i % 2 == 0 else PAPER)
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=rr, column=ci, value=v)
        cell.border = border; cell.alignment = center if ci in (1, 5) else left
        cell.fill = PatternFill("solid", fgColor=GREEN) if ci in entry_cols else band
        if ci == 6 and v:
            cell.font = Font(bold=True, color=INK)
    ws.row_dimensions[rr].height = 30

for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:R{len(rows)+2}"

# summary
s = wb.create_sheet("Summary")
named = sum(1 for r in rows if r.get("dm_name"))
li = sum(1 for r in rows if r.get("linkedin"))
eml = sum(1 for r in rows if em(r))
web = sum(1 for r in rows if r.get("website"))
cin = sum(1 for r in rows if r.get("cin"))
s["A1"] = "Summary"; s["A1"].font = Font(bold=True, size=14, color=INK)
for i, (k, v) in enumerate([
    ("Leads with phone", len(rows)),
    ("…also with named decision-maker", named),
    ("…also with LinkedIn", li),
    ("…also with email", eml),
    ("…also with website", web),
    ("…also with CIN", cin),
    ("Segment", "Startup India · Delhi-NCR × DPIIT-recognized × Scaling stage"),
    ("Phone source", "Company's own website / verified listing (accuracy-gated)"),
    ("Person/CIN source", "MCA / Tofler / Zauba registry, source-verified"),
], 3):
    s.cell(row=i, column=1, value=k).font = Font(bold=True, color=INK)
    s.cell(row=i, column=2, value=v)
s.column_dimensions["A"].width = 34; s.column_dimensions["B"].width = 56

wb.save(OUT)
print(f"Saved {OUT}")
print(f"  {len(rows)} phone leads | named DM {named} | LinkedIn {li} | email {eml} | website {web} | CIN {cin}")
